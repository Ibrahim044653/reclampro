"""Export du registre des réclamations (FR053, BR006).

Trois formats au choix :
- CSV : pratique pour Excel rapide
- XLSX : format officiel Excel formaté (entêtes, couleurs, largeurs)
- PDF : registre imprimable, pratique pour les inspections sur place
"""
import csv
import io
from datetime import datetime
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import select

from .. import models
from ..database import get_db
from .auth import utilisateur_admin_download

router = APIRouter(prefix="/api/exports", tags=["exports"])

COLONNES = [
    ("Code dossier", lambda r: r.code),
    ("Canal", lambda r: r.canal),
    ("Statut", lambda r: r.statut),
    ("Catégorie", lambda r: r.categorie),
    ("Sous-catégorie", lambda r: r.sous_categorie or ""),
    ("Priorité", lambda r: r.priorite),
    ("Reçue le", lambda r: r.date_reception.strftime("%d/%m/%Y %H:%M")),
    ("Échéance SLA", lambda r: r.date_echeance_sla.strftime("%d/%m/%Y %H:%M")),
    ("Clôturée le", lambda r: r.date_cloture.strftime("%d/%m/%Y %H:%M") if r.date_cloture else ""),
    ("Motif clôture", lambda r: r.motif_cloture or ""),
    ("Montant FCFA", lambda r: f"{r.montant_enjeu:,.0f}".replace(",", " ")),
    ("Client nom", lambda r: r.client.nom),
    ("Client prénom", lambda r: r.client.prenom),
    ("Client email", lambda r: r.client.email or ""),
    ("Client téléphone", lambda r: r.client.telephone or ""),
    ("Description", lambda r: (r.description or "").replace("\n", " ")[:500]),
]


def _toutes_reclamations(db: Session):
    return list(db.scalars(
        select(models.Reclamation).order_by(models.Reclamation.date_reception.desc())
    ))


@router.get("/registre.csv")
def export_csv(
    db: Session = Depends(get_db),
    _admin: models.Agent = Depends(utilisateur_admin_download),
):
    reclamations = _toutes_reclamations(db)
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=";")
    writer.writerow([c[0] for c in COLONNES])
    for r in reclamations:
        writer.writerow([fn(r) for _, fn in COLONNES])
    buffer.seek(0)
    nom_fichier = f"registre_reclamations_{datetime.utcnow():%Y%m%d}.csv"
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{nom_fichier}"'},
    )


@router.get("/registre.xlsx")
def export_xlsx(
    db: Session = Depends(get_db),
    _admin: models.Agent = Depends(utilisateur_admin_download),
):
    """Excel formaté pour les inspections BCEAO/CIMA (FR053)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    reclamations = _toutes_reclamations(db)
    wb = Workbook()
    ws = wb.active
    ws.title = "Registre des réclamations"

    # En-tête institutionnel
    ws.merge_cells("A1:H1")
    ws["A1"] = "Registre des réclamations clients"
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="185FA5")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = (
        f"Établissement : RéclamPro — Période d'extraction : {datetime.utcnow():%d/%m/%Y %H:%M} "
        f"— Total : {len(reclamations)} dossier(s)"
    )
    ws["A2"].font = Font(italic=True, size=10)
    ws["A2"].alignment = Alignment(horizontal="center")

    # Ligne d'en-tête colonnes
    header_row = 4
    header_fill = PatternFill("solid", fgColor="E6F1FB")
    header_font = Font(bold=True, color="0C447C")
    thin = Side(border_style="thin", color="B5D4F4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (label, _) in enumerate(COLONNES, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 30

    # Lignes de données
    for row_idx, r in enumerate(reclamations, start=header_row + 1):
        for col_idx, (_, fn) in enumerate(COLONNES, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=fn(r))
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        # Couleur de la ligne selon priorité
        if r.priorite == "CRITIQUE":
            for col_idx in range(1, len(COLONNES) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill("solid", fgColor="FCEBEB")
        elif r.priorite == "URGENT":
            for col_idx in range(1, len(COLONNES) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill("solid", fgColor="FAEEDA")

    # Largeurs
    largeurs = [20, 12, 14, 14, 22, 12, 18, 18, 18, 14, 14, 16, 16, 24, 18, 50]
    for i, w in enumerate(largeurs, start=1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)].width = w

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nom_fichier = f"registre_reclamations_{datetime.utcnow():%Y%m%d}.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nom_fichier}"'},
    )


@router.get("/rapport-mensuel.pdf")
def export_rapport_mensuel(
    annee: int,
    mois: int,
    db: Session = Depends(get_db),
    _admin: models.Agent = Depends(utilisateur_admin_download),
):
    """Rapport mensuel BCEAO/CIMA (RG012) au format PDF."""
    if not 1 <= mois <= 12:
        from fastapi import HTTPException
        raise HTTPException(422, "Mois invalide (1..12).")
    from ..services import rapport_mensuel
    contenu = rapport_mensuel.generer_pdf(db, annee, mois)
    nom_fichier = f"rapport_mensuel_{annee}-{mois:02d}.pdf"
    return StreamingResponse(
        iter([contenu]),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{nom_fichier}"'},
    )


@router.get("/registre.pdf")
def export_pdf(
    db: Session = Depends(get_db),
    _admin: models.Agent = Depends(utilisateur_admin_download),
):
    """Registre PDF pour archivage / inspection physique."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak,
    )

    reclamations = _toutes_reclamations(db)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=10 * mm, rightMargin=10 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    titre_style = ParagraphStyle(
        "titre", parent=styles["Title"], fontSize=16, textColor=colors.HexColor("#185FA5"),
    )
    meta_style = ParagraphStyle(
        "meta", parent=styles["Normal"], fontSize=9, textColor=colors.grey,
    )

    elements = [
        Paragraph("Registre des réclamations clients", titre_style),
        Paragraph(
            f"Établissement : RéclamPro — Extraction : {datetime.utcnow():%d/%m/%Y %H:%M} "
            f"— Total : <b>{len(reclamations)}</b> dossier(s)",
            meta_style,
        ),
        Spacer(1, 6 * mm),
    ]

    # On limite à 8 colonnes pour rester lisible en A4 paysage
    colonnes_pdf = [
        ("Code", lambda r: r.code),
        ("Canal", lambda r: r.canal),
        ("Statut", lambda r: r.statut),
        ("Catégorie", lambda r: r.categorie),
        ("Priorité", lambda r: r.priorite),
        ("Reçue", lambda r: r.date_reception.strftime("%d/%m/%y")),
        ("Échéance", lambda r: r.date_echeance_sla.strftime("%d/%m/%y")),
        ("Clôt.", lambda r: r.date_cloture.strftime("%d/%m/%y") if r.date_cloture else "—"),
        ("Motif", lambda r: r.motif_cloture or "—"),
        ("Client", lambda r: f"{r.client.prenom} {r.client.nom}"),
        ("Montant", lambda r: f"{r.montant_enjeu:,.0f}".replace(",", " ")),
    ]

    data = [[c[0] for c in colonnes_pdf]]
    for r in reclamations:
        data.append([str(fn(r)) for _, fn in colonnes_pdf])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#185FA5")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F5EE")]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#B5D4F4")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 8 * mm))
    elements.append(Paragraph(
        "Document généré automatiquement à des fins d'archivage et de présentation aux régulateurs "
        "(BCEAO / CIMA). Toute reproduction nécessite l'autorisation de l'établissement.",
        meta_style,
    ))

    doc.build(elements)
    buf.seek(0)
    nom_fichier = f"registre_reclamations_{datetime.utcnow():%Y%m%d}.pdf"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{nom_fichier}"'},
    )
